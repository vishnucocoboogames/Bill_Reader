import re
from pathlib import Path
from typing import Dict, List, Any
import openpyxl
import logging

logger = logging.getLogger(__name__)

class FileParserService:
    # Matches [ConsumerNumber]_[Month].xlsm or .xlsx etc
    FILENAME_PATTERN = re.compile(r"^([^_]+)_(.+)\.xlsx?m?$")

    def __init__(self, expected_fields: List[str] = None):
        # We expect these keys from 'Bill Parameter Mapping'
        self.expected_fields = expected_fields or [
            "Cno", "Cname", "Caddress", "MthYr", "CntdLoad"
        ]

    def map_files_by_consumer(self, directory: str) -> Dict[str, str]:
        """
        Regex-parses filenames to extract ConsumerNumber.
        Returns a map of ConsumerNumber -> absolute file path.
        """
        file_map = {}
        dir_path = Path(directory)
        
        if not dir_path.exists() or not dir_path.is_dir():
            logger.error(f"Directory not found: {directory}")
            return file_map

        # Allow .xlsx, .xls, .xlsm
        for file_path in dir_path.glob("*.xls*"):
            match = self.FILENAME_PATTERN.match(file_path.name)
            if match:
                consumer_number = match.group(1)
                file_map[consumer_number] = str(file_path.absolute())
            else:
                logger.warning(f"File {file_path.name} didn't match pattern [ConsumerNumber]_[Month].xlsm")
                
        return file_map

    def parse_excel(self, file_path: str) -> Dict[str, Any]:
        """
        Loads an excel file using openpyxl and extracts specific mappings.
        Evaluates formulas to get the actual text/numbers using data_only=True.
        """
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            data = {}
            
            # Parse Static Fields from 'Bill Parameter Mapping'
            if "Bill Parameter Mapping" in wb.sheetnames:
                mapping_sheet = wb["Bill Parameter Mapping"]
                
                # Extract specific meter config parameters from Table 1: column D
                # Only check user-specified static rows
                for r in [5, 13, 64, 66, 67, 68, 121, 125, 126]:
                    val = mapping_sheet.cell(row=r, column=4).value    # Column D
                    data[f"cfg_R{r}"] = val

                # Extract simple static code-value pairs from Table 3: columns J, K and L
                # J=Parameter (10), K=Code (11), L=Mapping Bill Value (12)
                for row in mapping_sheet.iter_rows(min_row=1, max_row=200, min_col=10, max_col=12):
                    param = row[0].value
                    code = row[1].value
                    val = row[2].value
                    if code:
                        data[str(code).strip()] = val
                    elif param:
                        data[str(param).strip()] = val
                        
                # Extract the actual reading metrics which are un-cached on the front sheet
                # Column F (6): "Curr. Month Cumulative Energy for TOU[X]" -> Column H (8) is FR
                # Column J (10): "Last Month Cumulative Energy for TOU[X]" -> Column L (12) is IR
                for r in range(1, min(200, mapping_sheet.max_row + 1)):
                    curr_label = mapping_sheet.cell(row=r, column=6).value
                    if curr_label and isinstance(curr_label, str):
                        lbl = curr_label.strip().upper()
                        # Strictly match Current Month Cumulative Energy while avoiding Apparent and Export rows
                        if "CURR" in lbl and "MONTH" in lbl and "CUMULATIVE" in lbl and "ENERGY" in lbl and "APPARENT" not in lbl and "EXPORT" not in lbl:
                            if "TOU1" in lbl:
                                data["Normal -FR"] = mapping_sheet.cell(row=r, column=8).value
                            elif "TOU2" in lbl:
                                data["Peak-FR"] = mapping_sheet.cell(row=r, column=8).value
                            elif "TOU3" in lbl:
                                data["Offpeak-FR"] = mapping_sheet.cell(row=r, column=8).value
                                
                    last_label = mapping_sheet.cell(row=r, column=10).value
                    if last_label and isinstance(last_label, str):
                        lbl = last_label.strip().upper()
                        # Strictly match Last Month Cumulative Energy while avoiding Apparent and Export rows
                        if "LAST" in lbl and "MONTH" in lbl and "CUMULATIVE" in lbl and "ENERGY" in lbl and "APPARENT" not in lbl and "EXPORT" not in lbl:
                            if "TOU1" in lbl:
                                data["Normal-IR"] = mapping_sheet.cell(row=r, column=12).value
                            elif "TOU2" in lbl:
                                data["Peak-IR"] = mapping_sheet.cell(row=r, column=12).value
                            elif "TOU3" in lbl:
                                data["Offpeak-IR"] = mapping_sheet.cell(row=r, column=12).value
            else:
                raise ValueError("Corrupted Excel structure. Missing 'Bill Parameter Mapping' sheet.")
                
            # Parse Readings from 'BILLDATA' (fallback for other fields)
            if "BILLDATA" in wb.sheetnames:
                bill_sheet = wb["BILLDATA"]
                for row in bill_sheet.iter_rows(min_row=1, max_row=200, min_col=4, max_col=7):
                    param = row[0].value
                    val = row[3].value
                    if param and val is not None:
                        data[str(param).strip()] = val


            
            # Validation Layer for schema
            missing_cols = [col for col in self.expected_fields if col not in data]
            if missing_cols:
                raise ValueError(f"Corrupted Excel structure. Missing expected fields: {missing_cols}")
                
            return data
            
        except Exception as e:
            logger.error(f"Error parsing Excel file {file_path}: {e}")
            raise
